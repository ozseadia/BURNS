# -*- coding: utf-8 -*-
"""
Created on Thu Apr 21 19:11:27 2022

@author: OzSea
"""
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os

class MetaData():
    
    
    def __init__(self,images_path):
        ImgesMetaDataExel=os.path.join(images_path,'MetaData.xlsx')
        MetaWB=Workbook()
        self.excelpath=ImgesMetaDataExel
        if not (os.path.exists(ImgesMetaDataExel)):
            page = MetaWB.active
            page.title = 'Images Meta Data'
            headers=['File Name','Size','Rank','Folder Name','Query','URL',
                     'PDF file name','Artical title','Author','is Human','is child',
                     'Burn','Abuse Burn']
            page.append(headers) # write the headers to the first line
            MetaWB.save(ImgesMetaDataExel)
            
            self.MetaWB=load_workbook(ImgesMetaDataExel)
            
        else:
            self.MetaWB = load_workbook(ImgesMetaDataExel)

    def MetaDataAppend(self,data):
        page = self.MetaWB.active
        page.append(data)
        self.MetaWB.save(self.excelpath)
        
        
        